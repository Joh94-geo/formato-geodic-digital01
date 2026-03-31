
import streamlit as st
import cv2
import numpy as np
from paddleocr import PaddleOCR
from openpyxl import load_workbook
import json
from io import BytesIO

st.set_page_config(page_title="Digitalizador de Formatos GEODIC", layout="wide")

# -----------------------------
# Cargar configuración
# -----------------------------
with open("config.json", "r") as f:
    config = json.load(f)

# -----------------------------
# Inicializar OCR manuscrito
# -----------------------------
ocr = PaddleOCR(use_angle_cls=True, lang="en")  # PaddleOCR con modelo inglés (el más estable para manuscrita)

# -----------------------------
# Funciones auxiliares
# -----------------------------
def crop_zone(img, zone):
    x1, y1, x2, y2 = zone
    return img[y1:y2, x1:x2]

def read_field(img, zone):
    crop = crop_zone(img, zone)
    result = ocr.ocr(crop, cls=True)
    if not result:
        return ""
    text = " ".join([line[1][0] for line in result[0]])
    return text.strip()

# -----------------------------
# Interfaz Streamlit
# -----------------------------
st.title("📄 Digitalizador Automático de Formatos – GEODIC")
st.write("Convierte tus formatos manuscritos en Excel automáticamente.")

uploaded_img = st.file_uploader("Sube la foto del formato diligenciado", type=["jpg", "jpeg", "png"])
uploaded_excel = st.file_uploader("Sube el archivo Excel base", type=["xlsx"])

if uploaded_img and uploaded_excel:
    # Leer imagen
    img_bytes = np.asarray(bytearray(uploaded_img.read()), dtype=np.uint8)
    img = cv2.imdecode(img_bytes, cv2.IMREAD_COLOR)

    st.image(img, caption="Imagen cargada", use_column_width=True)

    # Leer Excel base
    wb = load_workbook(uploaded_excel)
    ws = wb.active

    st.subheader("✅ Extrayendo campos manuscritos...")

    # -----------------------------
    # Procesar campos simples
    # -----------------------------
    for fname, fdata in config["simple_fields"].items():
        zone = fdata["zone"]
        cell = fdata["cell"]

        value = read_field(img, zone)
        ws[cell] = value

        st.write(f"**{fname}:** {value}")

    # -----------------------------
    # Procesar tabla
    # -----------------------------
    st.subheader("✅ Procesando filas de la tabla...")

    start_row = config["table"]["start_row"]
    col = config["table"]["column"]

    for i, zone in enumerate(config["table"]["zones"]):
        row = start_row + i
        value = read_field(img, zone)
        ws[f"{col}{row}"] = value
        st.write(f"Fila {row}: {value}")

    # -----------------------------
    # Generar archivo final
    # -----------------------------
    output = BytesIO()
    wb.save(output)

    st.success("✅ Formato diligenciado procesado correctamente")

    st.download_button(
        label="⬇️ Descargar Excel diligenciado",
        data=output.getvalue(),
        file_name="formato_diligenciado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
